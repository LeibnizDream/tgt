import os
import sys
import json
import string
import os
import shutil
import logging
import subprocess
import time
import urllib.request
import zipfile
import re

def get_materials_path(filename):
    """Get the path to a file in the materials directory."""
    parent_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        base_path = os.path.join(sys._MEIPASS, parent_dir, 'materials')
    except:
        base_path = os.path.join(os.getcwd(), parent_dir, 'materials')

    return os.path.join(base_path, filename)

def load_json_file(file_path):
    """Utility function to load JSON files with error handling."""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return json.load(file)
    except FileNotFoundError:
        print(f"Error: {file_path} not found.")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"Error: Failed to parse JSON from {file_path}.")
        sys.exit(1)

def load_text_file(filename):
    """Utility function to load text files with error handling."""
    # Resolve the file path using get_materials_path
    file_path = get_materials_path(filename)
    
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read().splitlines()
    except FileNotFoundError:
        print(f"Error: {file_path} not found.")
        sys.exit(1)

def set_global_variables():
    """Loads necessary configurations and paths."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    
    languages_path = os.path.join(parent_dir, 'materials', 'global_variables', 'LANGUAGES')
    columns_path = os.path.join(parent_dir, 'materials', 'global_variables', 'OBLIGATORY_COLUMNS')
    nolatin_path = os.path.join(parent_dir, 'materials', 'global_variables', 'NO_LATIN')

    LANGUAGES = load_json_file(languages_path)
    OBLIGATORY_COLUMNS = load_text_file(columns_path)
    NO_LATIN = load_text_file(nolatin_path)

    return LANGUAGES, NO_LATIN, OBLIGATORY_COLUMNS

def load_glossing_rules(filename):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)

    file = os.path.join(parent_dir, 'materials', 'glossing_rules', filename)

    return load_json_file(file)

def find_language(language, LANGUAGES):
    """Finds the language code by its name."""
    language_lower = language.lower()
    
    # Reverse the LANGUAGES dictionary for language name -> code lookup
    reversed_languages = {value.lower(): key for key, value in LANGUAGES.items()}
    
    language_code = reversed_languages.get(language_lower)
    
    if language_code:
        print(f'Language recognized: {language_code}')
        return language_code
    else:
        print(f"Unsupported language: {language}")
        sys.exit(1)

def clean_german_transcription(input_string: str) -> str:
    """
    Lowercase & strip punctuation from input_string,
    but leave any [LABEL: …] spans exactly as-is.
    """
    # Prepare your punctuation translator
    translator = str.maketrans("", "", string.punctuation)
    
    # This regex captures any [...] span as its own piece
    parts = re.split(r"(\[[^\]]+\])", input_string)
    
    cleaned_parts = []
    for part in parts:
        # If it’s an annotation (e.g. “[PER: Hans]”), leave untouched
        if re.fullmatch(r"\[[^\]]+\]", part):
            cleaned_parts.append(part)
        else:
            # Otherwise lowercase & remove punctuation
            cleaned_parts.append(part.lower().translate(translator))
    
    return "".join(cleaned_parts)


def install_ffmpeg():
    """
    Downloads FFmpeg, unpacks zip-file, deletes zip-path
    and installs FFmpeg to specific path
    """
    destination_path = os.path.expanduser("~")
    ffmpeg_url = "https://www.gyan.dev/ffmpeg/builds/ffmpeg-release-essentials.zip"

    zip_path = os.path.join(destination_path, "ffmpeg-7.1-essentials_build.zip")
    ffmpeg_extract_path = os.path.join(destination_path, "ffmpeg")

    try:
        print("Downloading ffmpeg...")
        urllib.request.urlretrieve(ffmpeg_url, zip_path)
        print("Download complete.")

        print(f"Extracting ffmpeg to {ffmpeg_extract_path}...")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(ffmpeg_extract_path)
        print("Extraction complete.")

        os.remove(zip_path)

        print(f"ffmpeg has been installed to {ffmpeg_extract_path}.")
        print("Adding path to system's PATH environment variable.")
        ffmpeg_path = os.path.join(ffmpeg_extract_path, "ffmpeg-7.1-essentials_build/bin/ffmpeg.exe")
        os.environ["PATH"] += os.pathsep + os.path.dirname(ffmpeg_path)
        return ffmpeg_path
    except Exception as e:
        print("An error occurred:", e)


def find_ffmpeg():
    """Dynamically finds ffmpeg executable path"""
    ffmpeg_path = shutil.which("ffmpeg")

    if not ffmpeg_path:
        print("FFmpeg not found. Attempting to install FFmpeg...")
        ffmpeg_path = install_ffmpeg()
    else:
        return ffmpeg_path


def format_excel_output(excel_output_file, columns_to_highlight: list, row_indices: set | None = None):
    import openpyxl
    from openpyxl.styles import Font
    from copy import copy

    wb = openpyxl.load_workbook(excel_output_file)
    ws = wb.active

    # Get header mapping
    headers = [cell.value for cell in ws[1]]
    col_indices = [i for i, h in enumerate(headers) if h in columns_to_highlight]

    # Apply red color to specified columns (only row_indices rows if provided)
    for row in ws.iter_rows(min_row=2):
        pandas_idx = row[0].row - 2  # Excel row 2 → pandas index 0
        if row_indices is not None and pandas_idx not in row_indices:
            continue
        for col_idx in col_indices:
            cell = row[col_idx]
            if cell.value:
                # Create a completely new Font object with all properties copied
                old_font = cell.font
                cell.font = Font(
                    name=old_font.name,
                    size=old_font.size,
                    bold=old_font.bold,
                    italic=old_font.italic,
                    vertAlign=old_font.vertAlign,
                    underline=old_font.underline,
                    strike=old_font.strike,
                    color="FF0000"  # Only this changes
                )

    wb.save(excel_output_file)


def ensure_ollama_running(host: str = "http://127.0.0.1:11434", timeout: int = 60) -> None:
    import sys

    def is_ready() -> tuple:
        try:
            resp = urllib.request.urlopen(host, timeout=2)
            return True, f"HTTP {resp.status}"
        except urllib.error.URLError as e:
            return False, str(e.reason)
        except Exception as e:
            return False, str(e)

    ready, status = is_ready()
    if ready:
        print(f"[Ollama] Already running at {host} ({status})", file=sys.stderr)
        return

    print(f"[Ollama] Not reachable at {host} — reason: {status}", file=sys.stderr)
    print("[Ollama] Launching 'ollama serve'...", file=sys.stderr)

    kwargs = {}
    if sys.platform == "win32":
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW

    try:
        proc = subprocess.Popen(
            ["ollama", "serve"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            **kwargs,
        )
        print(f"[Ollama] Process started with PID {proc.pid}", file=sys.stderr)
    except FileNotFoundError:
        raise RuntimeError("[Ollama] 'ollama' executable not found in PATH. Is Ollama installed?")
    except Exception as e:
        raise RuntimeError(f"[Ollama] Failed to launch process: {e}")

    deadline = time.time() + timeout
    attempt = 0
    while time.time() < deadline:
        attempt += 1
        time.sleep(2)
        ready, status = is_ready()
        print(f"[Ollama] Attempt {attempt} — {status}", file=sys.stderr)
        if ready:
            print(f"[Ollama] Ready after {attempt} attempts.", file=sys.stderr)
            return

        if proc.poll() is not None:
            out, err = proc.communicate()
            raise RuntimeError(
                f"[Ollama] Process exited early (code {proc.returncode}).\n"
                f"stdout: {out.decode(errors='replace')}\n"
                f"stderr: {err.decode(errors='replace')}"
            )

    out, err = proc.communicate(timeout=2)
    raise RuntimeError(
        f"[Ollama] Did not become ready within {timeout}s.\n"
        f"stdout: {out.decode(errors='replace')}\n"
        f"stderr: {err.decode(errors='replace')}"
    )